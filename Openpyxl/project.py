from typing import final
import openpyxl
from openpyxl.styles import Font

master_data = openpyxl.load_workbook('master_sheet.xlsx')
daily_data = openpyxl.load_workbook('daily_sheet.xlsx')

master_sheet = master_data['data']
daily_sheet = daily_data['Sheet1']

# Get row count for daily sheet
is_data = True
daily_row_count = 1

while is_data:
  daily_row_count += 1
  data = daily_sheet.cell(row=daily_row_count, column=1).value
  if data == None:
    is_data = False


# Get row count for master sheet
is_data = True
master_row_count = 1

while is_data:
  master_row_count += 1
  data = master_sheet.cell(row=master_row_count, column=1).value
  if data == None:
    is_data = False

# Get data from daily sheet
# Extract data --> store it into list of dictionaries
todays_data = []

for i in range(1, daily_row_count):
  row_data = {}
  row_data['id'] = daily_sheet.cell(row=i, column=1).value
  row_data['todays_purchase'] = daily_sheet.cell(row=i, column=2).value
  row_data['todays_rewards'] = daily_sheet.cell(row=i, column=3).value
  todays_data.append(row_data)

#print(todays_data)

#[{'id': 'ID', 'todays_purchase': 'Totdays purchases', 'todays_rewards': 'Todays Reward'}]
# Write daily sheet data into master excel sheet
# Find row using the ITD
# Go to total pruchase cell + today purchase
# Go to total reward balance + todays reward

for i in range(2, master_row_count):
  id = master_sheet.cell(row=i, column=1).value
  for row in todays_data:
    if row['id'] == id:
      todays_purchase = int(row['todays_purchase'])
      todays_reward = int(row['todays_rewards'])

      # Get data from master sheet
      total_purchase = master_sheet.cell(row=i, column=6).value
      total_reward = master_sheet.cell(row=i, column=7).value

      #Add values of todays purchase into total data
      new_total_purchase = total_purchase + todays_purchase
      new_total_reward = total_reward + todays_reward

      master_sheet.cell(row=i, column=6).value = new_total_purchase
      master_sheet.cell(row=i, column=7).value = new_total_reward

master_data.save('master_sheet.xlsx')

daily_report = openpyxl.Workbook()
ws = daily_report.active

# Get headers
is_data = True
column_count = 1
header_values = []

while is_data:
  column_count += 1
  data = master_sheet.cell(row=1, column=column_count).value
  if data != None:
    header_values.append(data)
  else:
    is_data = False

header_style = Font(name="Times New Roman", size=12, bold=True)

for i, col_name in enumerate(header_values):
  col_index = i + 1
  ws.cell(row=1, column=col_index).value = col_name
  ws.cell(row=1, column=col_index).font = header_style

IDs = []
for data in todays_data:
  IDs.append(data['id'])
#remove the first cell from list which is ID header
IDs.pop(0)

final_data = []
for i in range(2,master_row_count):
  id = master_sheet.cell(row=i, column=1).value
  if id in IDs:
    lst = []
    for j in range(2,8):
      lst.append(master_sheet.cell(row=i, column=j).value)
    final_data.append(lst)

for data in final_data:
  ws.append(data)

daily_report.save("daily_report_send.xlsx")