import openpyxl
from openpyxl.styles import PatternFill

#link absolute file path
wb = openpyxl.load_workbook("balance.xlsx")
ws = wb['Sheet1']

fill_pattern = PatternFill(patternType='solid',
                               fgColor='C64747')  

ws['B4'].fill = fill_pattern

wb.save("Balance.xlsx")