import openpyxl
from openpyxl.styles import Font

#link absolute file path
wb = openpyxl.load_workbook("balance.xlsx")
ws = wb['Score']

ws['B11'] = "=SUM(B2:B9)"
ws['B12'] = "=AVERAGE(B2:B9)"

# (Balance * interest) * balance

ws['D1'] = 'Balance after a year'
ws['D1'].font = Font(bold=True,name='Arial',size=10)

for i in range(2,9):
  balance = ws.cell(row=i, column=2).value
  interest = ws.cell(row=i, column=3).value
  final_balance = (balance * interest) + balance
  ws.cell(row=i,column=4).value = final_balance

wb.save("balance.xlsx")