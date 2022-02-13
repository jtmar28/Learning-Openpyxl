import openpyxl

#link absolute file path
wb = openpyxl.load_workbook("balance.xlsx")
ws = wb['Sheet1']

print(ws['BB'].value)

#ws['A9'].value = "Rick"
#ws['B9'].value = 1500

#ws.cell(row=5,column=2).value = 50000

#ws['C1'] = 'Double Balance'

for i in range(2,10):
  b_col = ws.cell(row=i,column=2).value
  c_value = b_col * 2
  ws.cell(row=i, column = 3).value = c_value

wb.save("balanxe.xlsx")