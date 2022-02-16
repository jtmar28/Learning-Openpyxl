import openpyxl
from openpyxl.styles import numbers

#link absolute file path
wb = openpyxl.load_workbook("balance.xlsx")
ws = wb['Sheet1']

ws['C4'] = '11/11/20'
ws['C4'] = numbers.FORMAT_DATE_DATETIME

ws['D4'] = 20
ws['E4'] = 'Beginner'
ws['E4'].number_format = numbers.FORMAT_TEXT

wb.save("balance.xlsx")