import openpyxl

#link absolute file path
wb = openpyxl.load_workbook("balance.xlsx")
ws = wb['Sheet1']

#created a generator object
rows = ws.iter_rows(min_row=1,max_row=7,min_col=1,max_col=2)

names = []
balance = []
for a,b in rows:
  names.append(a.value)
  balance.append(b.value)

print(names)
print(balance)

columns = ws.iter_cols(min_row=1,max_row=5,min_col=1,max_col=2)
#Tuple for each column
for col in columns:
  print(col)

#All rows and all columns
rows = list(ws.rows)
columns = list(ws.column)
