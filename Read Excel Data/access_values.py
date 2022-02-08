import openpyxl

#link absolute file path
wb = openpyxl.load_workbook("balance.xlsx")
ws = wb['Sheet1']

# Single cell value
print(ws['B5'].value)

#Single cell value but specifiying 
print(ws.cell(row=6, column=1).value)

#this is printed as a tuple
value_range = ws['A2':'B5']

for a,b in value_range:
  print(a.value, b.value)