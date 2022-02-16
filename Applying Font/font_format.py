import openpyxl
from openpyxl import Font, Color

#link absolute file path
wb = openpyxl.load_workbook("balance.xlsx")
ws = wb['Sheet1']

# font_style = Font(name="Chalkboard",size=14,color="#0D8377", italic=True,
#                   bold=True)

# a4 = ws['A4']
# a4.font = font_style

f_style = Font(name="Reem Kufi", size=12, color="DB3822",
    underline='single', strikethrough=True)

for i in range(2,10):
  ws.cell(row=i, column=3).font = f_style

wb.save("balanxe.xlsx")