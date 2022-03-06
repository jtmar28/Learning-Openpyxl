import openpyxl
import file_names

excel_files = file_names.files
values = []

#for file in excel_files:
#  workbook = openpyxl.load_workbook(file)
#  worksheet = workbook['Views']
#  cell_value = worksheet['C56'].value
#  values.append(cell_value)
#  print(cell_value)

start = 1
for file_num in excel_files:
  Feburary2022Usage = openpyxl.load_workbook('Feburary2022Usage.xlsx')
  workbook = openpyxl.load_workbook(file_num)
  worksheet = workbook['Views']
  cell_value1 = worksheet['C56'].value
  worksheet2 = Feburary2022Usage['Sheet1']
  start = start + 1
  cell_title1 = "Page"
  cell_title2 = "Usage"
  cell_value2 = "B" + str(start)
  worksheet2["A1"] = "Page"
  worksheet2["B1"] = "Usage"
  worksheet2[cell_value2] = cell_value1
  Feburary2022Usage.save("Feburary2022Usage.xlsx")

print("Successfuly Compiled.")
